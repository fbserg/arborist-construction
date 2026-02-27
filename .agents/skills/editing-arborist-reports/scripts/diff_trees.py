"""Diff tree data between a docx report and an updated Excel/TSV source.

Usage:
    python3 diff_trees.py <report.docx> <updated.xlsx>     # Excel mode
    python3 diff_trees.py <report.docx> --stdin             # TSV from stdin

Outputs [project]/.work/diff.json and prints a human-readable summary.
"""

import sys, os, json, csv, io

# Reuse extract_trees internals
sys.path.insert(0, os.path.dirname(__file__))
from extract_trees import parse_tree_table, _map_header, _coerce, _TREE_FIELD_MAP

# Report-relevant Excel columns (A–K)
_MAX_EXCEL_COL = 11


def _docx_trees(docx_path):
    """Extract 'before' tree data from docx via pandoc markdown."""
    import subprocess
    result = subprocess.run(
        ["pandoc", docx_path, "-t", "markdown"],
        capture_output=True, text=True, encoding='utf-8'
    )
    if result.returncode != 0:
        print(f"pandoc error: {result.stderr}", file=sys.stderr)
        sys.exit(1)
    return parse_tree_table(result.stdout)


def _read_excel(xlsx_path):
    """Read trees from Inventory sheet, columns A–K."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # Find inventory sheet (case-insensitive)
    sheet_name = None
    for name in wb.sheetnames:
        if name.lower() == 'inventory':
            sheet_name = name
            break
    if sheet_name is None:
        print(f"No 'Inventory' sheet in {xlsx_path}", file=sys.stderr)
        print(f"Available sheets: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)
    ws = wb[sheet_name]

    # Find header row (contains "TREE" in column A or B)
    header_row = None
    for r in range(1, min(ws.max_row + 1, 10)):
        for c in range(1, _MAX_EXCEL_COL + 1):
            v = ws.cell(r, c).value
            if v and 'tree' in str(v).lower():
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        print("Cannot find header row in Inventory sheet", file=sys.stderr)
        sys.exit(1)

    # Map headers
    headers = []
    for c in range(1, _MAX_EXCEL_COL + 1):
        v = ws.cell(header_row, c).value
        headers.append(str(v).strip() if v else '')

    field_map = {}  # col_index -> (field_name, type)
    for i, h in enumerate(headers):
        if not h:
            continue
        mapped = _map_header(h)
        if mapped:
            field_map[i] = mapped

    # Read data rows
    trees = []
    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c + 1).value for c in range(_MAX_EXCEL_COL)]
        # Skip empty rows
        if not any(v is not None and str(v).strip() for v in vals):
            continue
        tree = {}
        for i, (field, type_name) in field_map.items():
            raw = vals[i]
            if raw is None:
                continue
            val = _coerce(str(raw), type_name)
            if val is not None:
                tree[field] = val
        if tree:
            trees.append(tree)
    return trees


def _read_tsv(stream):
    """Read trees from tab-separated text (stdin)."""
    text = stream.read()
    reader = csv.DictReader(io.StringIO(text), delimiter='\t')
    field_map = {}  # csv_header -> (field_name, type)
    for h in reader.fieldnames or []:
        mapped = _map_header(h.strip())
        if mapped:
            field_map[h] = mapped

    trees = []
    for row in reader:
        tree = {}
        for h, (field, type_name) in field_map.items():
            raw = row.get(h, '').strip()
            if not raw:
                continue
            val = _coerce(raw, type_name)
            if val is not None:
                tree[field] = val
        if tree:
            trees.append(tree)
    return trees


def _normalize(val):
    """Normalize a value for comparison."""
    if val is None:
        return None
    if isinstance(val, float):
        # Round to 1 decimal to avoid Excel floating-point artifacts
        val = round(val, 1)
        if val == int(val):
            return int(val)
        return val
    if isinstance(val, str):
        # Normalize whitespace and non-breaking spaces
        return ' '.join(val.replace('\xa0', ' ').split())
    return val


def diff_trees(before, after):
    """Compare two tree lists. Returns (changes, added, removed)."""
    before_map = {}
    for t in before:
        tid = t.get('id')
        if tid is not None:
            before_map[tid] = t
    after_map = {}
    for t in after:
        tid = t.get('id')
        if tid is not None:
            after_map[tid] = t

    # Only compare fields present in the after data (handles partial TSV input)
    after_fields = {k for t in after for k in t if k != 'id'}
    all_fields = sorted(after_fields)

    changes = []
    for tid in sorted(set(before_map) & set(after_map)):
        b, a = before_map[tid], after_map[tid]
        diffs = {}
        for field in all_fields:
            old = _normalize(b.get(field))
            new = _normalize(a.get(field))
            if old != new:
                diffs[field] = {"old": old, "new": new}
        if diffs:
            changes.append({"tree_id": tid, "fields": diffs})

    added = sorted(set(after_map) - set(before_map))
    removed = sorted(set(before_map) - set(after_map))

    return changes, added, removed


def main():
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <report.docx> <updated.xlsx|--stdin>", file=sys.stderr)
        sys.exit(1)

    docx_path = sys.argv[1]
    mode = sys.argv[2]

    # Extract before
    before = _docx_trees(docx_path)
    if not before:
        print("Warning: no trees found in docx", file=sys.stderr)

    # Extract after
    if mode == '--stdin':
        after = _read_tsv(sys.stdin)
    else:
        after = _read_excel(mode)

    if not after:
        print("Warning: no trees found in source", file=sys.stderr)

    # Diff
    changes, added, removed = diff_trees(before, after)

    result = {
        "changes": changes,
        "added_trees": added,
        "removed_trees": removed,
    }

    # Write to .work/diff.json
    report_dir = os.path.dirname(os.path.abspath(docx_path))
    work_dir = os.path.join(report_dir, ".work")
    os.makedirs(work_dir, exist_ok=True)
    out_path = os.path.join(work_dir, "diff.json")
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    # Print summary
    if not changes and not added and not removed:
        print("No differences found.")
    else:
        for c in changes:
            tid = c["tree_id"]
            for field, vals in c["fields"].items():
                print(f"  Tree {tid}: {field}: {vals['old']!r} → {vals['new']!r}")
        for tid in added:
            print(f"  Tree {tid}: NEW (not in docx)")
        for tid in removed:
            print(f"  Tree {tid}: REMOVED (not in source)")
        print(f"\nWrote {out_path}")


if __name__ == '__main__':
    main()
